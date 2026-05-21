from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Table, TableStyle


DISTRICTS = [
    "GORAKHPUR",
    "DEORIA",
    "KUSHI NAGAR",
    "MAHRAJGANJ",
    "BASTI",
    "SANT KABEER NAGAR",
    "SIDDHARTH NAGAR",
    "GONDA",
    "BAHRAICH",
    "BALRAMPUR",
    "SHRAVASTI",
]

GKP_RANGE = ["GORAKHPUR", "DEORIA", "KUSHI NAGAR", "MAHRAJGANJ"]
BASTI_RANGE = ["BASTI", "SANT KABEER NAGAR", "SIDDHARTH NAGAR"]
DEVIPATAN_RANGE = ["GONDA", "BAHRAICH", "BALRAMPUR", "SHRAVASTI"]


class ReportGenerationError(Exception):
    """Raised when uploaded NCRP data cannot be converted into a report."""


def find_district_column(df):
    for col in df.columns:
        cname = str(col).lower()
        if "district" in cname or "police station" in cname:
            return col
    raise ReportGenerationError("District column not found in one of the CSV files.")


def find_column(df, keywords):
    for col in df.columns:
        cname = str(col).lower()
        if all(k in cname for k in keywords):
            return col
    raise ReportGenerationError(f"Required column not found: {' + '.join(keywords)}")


def normalize(df):
    dist = find_district_column(df)
    tc = find_column(df, ["total", "complaint"])
    amt = find_column(df, ["amount", "reported"])
    lien = find_column(df, ["lien", "amount"])

    out = df[[dist, tc, amt, lien]].copy()
    out.columns = ["District", "Total Complaint", "Amount Raw", "Lien Raw"]
    out["District"] = (
        out["District"]
        .astype(str)
        .str.upper()
        .str.replace("DISTRICT", "", regex=False)
        .str.strip()
    )
    out["Total Complaint"] = (
        pd.to_numeric(out["Total Complaint"], errors="coerce").fillna(0).astype(int)
    )
    out["Amount Raw"] = pd.to_numeric(out["Amount Raw"], errors="coerce").fillna(0)
    out["Lien Raw"] = pd.to_numeric(out["Lien Raw"], errors="coerce").fillna(0)

    max_val = max(out["Amount Raw"].max(), out["Lien Raw"].max())
    divisor = 100000 if max_val > 100000 else 1
    out["Amount Reported"] = out["Amount Raw"] / divisor
    out["Lien Amount"] = out["Lien Raw"] / divisor
    return out[["District", "Total Complaint", "Amount Reported", "Lien Amount"]]


def load_selected_files(cumulative_csv, daily_csv):
    try:
        cumulative = (
            normalize(pd.read_csv(cumulative_csv))
            .set_index("District")
            .reindex(DISTRICTS, fill_value=0)
        )
        daily = (
            normalize(pd.read_csv(daily_csv))
            .set_index("District")
            .reindex(DISTRICTS, fill_value=0)
        )
    except UnicodeDecodeError as exc:
        raise ReportGenerationError("CSV encoding is not readable. Please export as UTF-8 CSV.") from exc
    except pd.errors.EmptyDataError as exc:
        raise ReportGenerationError("One uploaded CSV file is empty.") from exc
    return cumulative, daily


def pct(numerator, denominator):
    return numerator / denominator if denominator else 0


def money(value):
    return f"₹ {value:.2f}"


def percent(value):
    return f"{value:.2%}"


def is_total_label(value):
    text = str(value).upper()
    return "RANGE" in text or "ZONE" in text


def build_report_df(cumulative, daily):
    def build_row(sn, district):
        p_tc, p_amt, p_lien = cumulative.loc[district]
        d_tc, d_amt, d_lien = daily.loc[district]
        t_tc = p_tc + d_tc
        t_amt = p_amt + d_amt
        t_lien = p_lien + d_lien
        return [
            sn,
            district,
            p_tc,
            p_amt,
            p_lien,
            pct(p_lien, p_amt),
            d_tc,
            d_amt,
            d_lien,
            pct(d_lien, d_amt),
            t_tc,
            t_amt,
            t_lien,
            pct(t_lien, t_amt),
        ]

    def build_total(label, districts):
        p = cumulative.loc[districts].sum()
        d = daily.loc[districts].sum()
        t = p + d
        return [
            label,
            "",
            p["Total Complaint"],
            p["Amount Reported"],
            p["Lien Amount"],
            pct(p["Lien Amount"], p["Amount Reported"]),
            d["Total Complaint"],
            d["Amount Reported"],
            d["Lien Amount"],
            pct(d["Lien Amount"], d["Amount Reported"]),
            t["Total Complaint"],
            t["Amount Reported"],
            t["Lien Amount"],
            pct(t["Lien Amount"], t["Amount Reported"]),
        ]

    rows = []
    sn = 1
    for district in GKP_RANGE:
        rows.append(build_row(sn, district))
        sn += 1
    rows.append(build_total("RANGE GORAKHPUR TOTAL", GKP_RANGE))

    for district in BASTI_RANGE:
        rows.append(build_row(sn, district))
        sn += 1
    rows.append(build_total("RANGE BASTI TOTAL", BASTI_RANGE))

    for district in DEVIPATAN_RANGE:
        rows.append(build_row(sn, district))
        sn += 1
    rows.append(build_total("RANGE DEVIPATAN TOTAL", DEVIPATAN_RANGE))
    rows.append(build_total("TOTAL GORAKHPUR ZONE", DISTRICTS))

    return pd.DataFrame(
        rows,
        columns=[
            "SNo",
            "District",
            "P_TC",
            "P_AMT",
            "P_LIEN",
            "P_PCT",
            "D_TC",
            "D_AMT",
            "D_LIEN",
            "D_PCT",
            "T_TC",
            "T_AMT",
            "T_LIEN",
            "T_PCT",
        ],
    )


def report_dates(generation_time=None):
    generation_time = generation_time or datetime.now()
    previous_end = generation_time - timedelta(days=1)
    previous_start = previous_end - timedelta(days=29)
    return {
        "today": generation_time.strftime("%d.%m.%Y"),
        "title_date": generation_time.strftime("%d.%m.%Y"),
        "title_time": generation_time.strftime("%H:%M"),
        "previous_period": (
            f"{previous_start.strftime('%d.%m.%Y')} TO {previous_end.strftime('%d.%m.%Y')}"
        ),
    }


def write_excel(df, out_xls, today, previous_period, title_date, title_time):
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    fills = {
        "title": PatternFill("solid", fgColor="F4B183"),
        "prev": PatternFill("solid", fgColor="FCE4D6"),
        "today": PatternFill("solid", fgColor="C6E0B4"),
        "total": PatternFill("solid", fgColor="DDEBF7"),
        "range": PatternFill("solid", fgColor="BDD7EE"),
    }
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:N2")
    ws["A1"] = f"Lien Amount GORAKHPUR ZONE (Date {title_date} & Time {title_time})"
    ws["A1"].font = Font(name="Calibri", bold=True, size=24)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = fills["title"]

    ws.merge_cells("A3:A4")
    ws["A3"] = "S No."
    ws.merge_cells("B3:B4")
    ws["B3"] = "District"
    ws.merge_cells("C3:F3")
    ws["C3"] = previous_period
    ws.merge_cells("G3:J3")
    ws["G3"] = today
    ws.merge_cells("K3:N3")
    ws["K3"] = "Total"

    headers = [
        "Total\nComplaint",
        "Amount\nReported\n(Rs. In Lacs)",
        "Lien\nAmount\n(Rs. In Lacs)",
        "Lien %",
    ] * 3
    for idx, header in enumerate(headers, start=3):
        ws.cell(row=4, column=idx, value=header)

    for row_idx, row in enumerate(df.itertuples(index=False), start=5):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=14):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.font = Font(name="Calibri", size=16)

    for row_num in (3, 4):
        for col in range(1, 15):
            cell = ws.cell(row=row_num, column=col)
            cell.font = Font(name="Calibri", size=16, bold=True)
            if col <= 6:
                cell.fill = fills["prev"]
            elif col <= 10:
                cell.fill = fills["today"]
            else:
                cell.fill = fills["total"]

    for cell in ws[1]:
        cell.fill = fills["title"]
        cell.font = Font(name="Calibri", bold=True, size=24)

    for row_idx in range(5, ws.max_row + 1):
        for col in range(3, 7):
            ws.cell(row=row_idx, column=col).fill = fills["prev"]
        for col in range(7, 11):
            ws.cell(row=row_idx, column=col).fill = fills["today"]
        for col in range(11, 15):
            ws.cell(row=row_idx, column=col).fill = fills["total"]

    for row_idx in range(5, ws.max_row + 1):
        label = str(ws.cell(row=row_idx, column=1).value).upper()
        if is_total_label(label):
            for col in range(1, 15):
                cell = ws.cell(row=row_idx, column=col)
                cell.font = Font(name="Calibri", size=16, bold=True)
                cell.fill = fills["title"] if "ZONE" in label else fills["range"]
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)

    for col in [6, 10, 14]:
        for row in range(5, ws.max_row + 1):
            ws.cell(row=row, column=col).number_format = "0.00%"
    for col in [4, 5, 8, 9, 12, 13]:
        for row in range(5, ws.max_row + 1):
            ws.cell(row=row, column=col).number_format = r"\₹\ #,##0.00"

    widths = {"A": 8.73, "B": 33.82, "C": 15.73}
    for col in range(3, 15):
        letter = get_column_letter(col)
        widths.setdefault(letter, 16 if letter in {"D", "E", "H", "I", "L", "M"} else 13)
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.row_dimensions[3].height = 21
    ws.row_dimensions[4].height = 63
    for row_idx in range(5, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 45.5

    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.2, right=0.2, top=0.25, bottom=0.25)
    wb.save(out_xls)


def register_pdf_fonts():
    font_candidates = [
        (
            "Arial",
            "Arial-Bold",
            Path(r"C:\Windows\Fonts\arial.ttf"),
            Path(r"C:\Windows\Fonts\arialbd.ttf"),
        ),
        (
            "DejaVuSans",
            "DejaVuSans-Bold",
            Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
            Path("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
        ),
    ]
    for regular_name, bold_name, regular_path, bold_path in font_candidates:
        if regular_path.exists() and bold_path.exists():
            pdfmetrics.registerFont(TTFont(regular_name, str(regular_path)))
            pdfmetrics.registerFont(TTFont(bold_name, str(bold_path)))
            return regular_name, bold_name
    return "Helvetica", "Helvetica-Bold"


def write_pdf(df, out_pdf, today, previous_period, title_date, title_time):
    font_name, bold_font_name = register_pdf_fonts()
    doc = SimpleDocTemplate(
        str(out_pdf),
        pagesize=landscape(A4),
        rightMargin=18,
        leftMargin=18,
        topMargin=18,
        bottomMargin=18,
    )
    header_para = ParagraphStyle(
        "header",
        fontName=bold_font_name,
        fontSize=9,
        leading=10.5,
        alignment=1,
    )
    title_para = ParagraphStyle(
        "title",
        fontName=bold_font_name,
        fontSize=18,
        leading=20,
        alignment=1,
    )

    def h(text):
        return Paragraph(str(text).replace(" ", "&nbsp;").replace("\n", "<br/>"), header_para)

    pdf_data = [
        [Paragraph(f"Lien Amount GORAKHPUR ZONE (Date {title_date} & Time {title_time})", title_para)] + [""] * 13,
        ["S No.", "District", previous_period, "", "", "", today, "", "", "", "Total", "", "", ""],
        [
            "",
            "",
            h("Total\nComplaint"),
            h("Amount\nReported\n(Rs. In Lacs)"),
            h("Lien\nAmount\n(Rs. In Lacs)"),
            h("Lien %"),
            h("Total\nComplaint"),
            h("Amount\nReported\n(Rs. In Lacs)"),
            h("Lien\nAmount\n(Rs. In Lacs)"),
            h("Lien %"),
            h("Total\nComplaint"),
            h("Amount\nReported\n(Rs. In Lacs)"),
            h("Lien\nAmount\n(Rs. In Lacs)"),
            h("Lien %"),
        ],
    ]

    for row in df.itertuples(index=False):
        pdf_data.append(
            [
                row.SNo,
                row.District,
                int(row.P_TC),
                money(row.P_AMT),
                money(row.P_LIEN),
                percent(row.P_PCT),
                int(row.D_TC),
                money(row.D_AMT),
                money(row.D_LIEN),
                percent(row.D_PCT),
                int(row.T_TC),
                money(row.T_AMT),
                money(row.T_LIEN),
                percent(row.T_PCT),
            ]
        )

    col_widths = [32.8, 105.0, 48, 62, 62, 38, 48, 62, 62, 38, 48, 62, 62, 38]
    row_heights = [34, 24, 56] + [24] * (len(pdf_data) - 3)
    table = Table(pdf_data, colWidths=col_widths, rowHeights=row_heights, repeatRows=3)
    style = TableStyle(
        [
            ("SPAN", (0, 0), (-1, 0)),
            ("SPAN", (0, 1), (0, 2)),
            ("SPAN", (1, 1), (1, 2)),
            ("SPAN", (2, 1), (5, 1)),
            ("SPAN", (6, 1), (9, 1)),
            ("SPAN", (10, 1), (13, 1)),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("FONTNAME", (0, 0), (-1, -1), font_name),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F4B183")),
            ("BACKGROUND", (0, 1), (5, 2), colors.HexColor("#FCE4D6")),
            ("BACKGROUND", (6, 1), (9, 2), colors.HexColor("#C6E0B4")),
            ("BACKGROUND", (10, 1), (13, 2), colors.HexColor("#DDEBF7")),
            ("BACKGROUND", (2, 3), (5, -1), colors.HexColor("#FCE4D6")),
            ("BACKGROUND", (6, 3), (9, -1), colors.HexColor("#C6E0B4")),
            ("BACKGROUND", (10, 3), (13, -1), colors.HexColor("#DDEBF7")),
            ("FONTNAME", (0, 0), (-1, 2), bold_font_name),
            ("LEFTPADDING", (0, 0), (-1, -1), 2),
            ("RIGHTPADDING", (0, 0), (-1, -1), 2),
            ("TOPPADDING", (0, 0), (-1, -1), 1),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
        ]
    )

    for i, row in enumerate(pdf_data):
        txt = str(row[0]).upper()
        if "ZONE" in txt:
            style.add("BACKGROUND", (0, i), (-1, i), colors.HexColor("#F4B183"))
            style.add("FONTNAME", (0, i), (-1, i), bold_font_name)
            if i >= 3:
                style.add("SPAN", (0, i), (1, i))
        elif "RANGE" in txt:
            style.add("BACKGROUND", (0, i), (-1, i), colors.HexColor("#BDD7EE"))
            style.add("FONTNAME", (0, i), (-1, i), bold_font_name)
            style.add("SPAN", (0, i), (1, i))

    table.setStyle(style)
    doc.build([table])


def generate_report(cumulative_csv, daily_csv, output_dir):
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    cumulative, daily = load_selected_files(cumulative_csv, daily_csv)
    report_df = build_report_df(cumulative, daily)
    dates = report_dates(datetime.now())
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")

    pdf_path = output_dir / f"Daily_Report_GKR_{stamp}.pdf"
    excel_path = output_dir / f"Daily_Report_GKR_{stamp}.xlsx"
    write_pdf(
        report_df,
        pdf_path,
        dates["today"],
        dates["previous_period"],
        dates["title_date"],
        dates["title_time"],
    )
    write_excel(
        report_df,
        excel_path,
        dates["today"],
        dates["previous_period"],
        dates["title_date"],
        dates["title_time"],
    )
    return {
        "pdf_path": pdf_path,
        "excel_path": excel_path,
        "title_date": dates["title_date"],
        "title_time": dates["title_time"],
        "previous_period": dates["previous_period"],
    }
